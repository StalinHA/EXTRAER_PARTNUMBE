import streamlit as st
import pandas as pd
import zipfile
import json
import re
from io import BytesIO
import base64
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import os
import tempfile
import shutil
from pathlib import Path

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(
    page_title="Dashboard de Fichas Técnicas",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- ESTILOS CSS PERSONALIZADOS ---
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1E3A8A;
        text-align: center;
        padding: 1rem 0;
        border-bottom: 3px solid #3B82F6;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #F3F4F6;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        border-left: 5px solid #3B82F6;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .metric-value {
        font-size: 2rem;
        font-weight: bold;
        color: #1E3A8A;
    }
    .metric-label {
        font-size: 0.9rem;
        color: #4B5563;
    }
    .stProgress > div > div > div > div {
        background-color: #3B82F6;
    }
    .success-box {
        padding: 1rem;
        background-color: #D1FAE5;
        border-radius: 0.5rem;
        border-left: 5px solid #10B981;
        margin: 1rem 0;
    }
    .info-box {
        padding: 1rem;
        background-color: #DBEAFE;
        border-radius: 0.5rem;
        border-left: 5px solid #3B82F6;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# --- DICCIONARIOS Y CONSTANTES ---
CATEGORIAS_OFICIALES = {
    '11743': 'COMPUTADORA PORTATIL',
    '11744': 'ESTACION DE TRABAJO PORTATIL',
    '11745': 'TABLETA',
    '11738': 'ESCANER DE DOCUMENTOS',
    '11735': 'COMPUTADORA DE ESCRITORIO',
    '11736': 'COMPUTADORA TODO EN UNO',
    '11740': 'ESTACION DE TRABAJO',
    '11741': 'MONITOR',
    '11742': 'PANTALLA PUBLICITARIA',
    '11747': 'DISPOSITIVOS DE ALMACENAMIENTO EXTERNO',
    '11749': 'PANTALLA INTERACTIVA',
    '11751': 'DISPOSITIVOS DE ALMACENAMIENTO INTERNO'
}

MARCAS_COMPLETAS = [
    'TRIUMPH BOARD', 'VASTEC', 'RHINOBOX', 'LENOVO', 'EXIN', 'M4X', 'KENYA TECHNOLOGY',
    'HP', 'MADI-TEK', 'INVESTMENT & BUSINESS SMART SBI', 'ADVANCE', 'QUI-TECH', 'TEXCOPER',
    'WIDETEK', 'IQTOUCH', 'LG', 'ONESCREEN', 'HUAWEI', 'INOTEC', 'DELL', 'I3', 'ASUS',
    'QUAMTU', 'KODAK', 'RICOH', 'SHARP', 'CIBER', 'HAO TECH', 'BROTHER', 'VIEWSONIC',
    'AVISION', 'SAMSUNG', 'ALLWIYA', 'GAMEMAX', 'DYNABOOK', 'HIPPOBOX', 'CONTEX', 'INNEX',
    'CTOUCH', 'HIKVISION', 'ZKT ECO', 'YEALINK', 'TEROS', 'SILVER VOLT', 'QOSOFT',
    'MIMIO', 'HAITECH', 'OPTOMA TECHNOLOGY INC', 'GROWTH HACK', 'MSI', 'XEROX', 'QOMO',
    'EPSON', 'CLEVERTOUCH', 'I2S INNOVATIVE IMAGING SOLUTIONS', 'IQ BOARD', 'GCS', 'COLORTRAC',
    'CANON', 'BOOKEYE', 'JFA TECHNOLOGY', 'AMC', 'MAXTIC', 'SANDISK', 'KINGSTON', 'ADATA',
    'NEW KRAL', 'TLC'
]

# --- FUNCIONES DE PROCESAMIENTO ---
def extract_part_number(descripcion):
    """
    Extrae el número de parte de la descripción usando patrones flexibles.
    Busca secuencias alfanuméricas con guiones o sin ellos, que suelen ser el part number.
    """
    if not descripcion or not isinstance(descripcion, str):
        return "No especificado"

    # Patrón mejorado para buscar números de parte
    patrones = [
        r'(?:UNIDAD\s+)?([A-Z0-9]{8,}(?:-[A-Z0-9]+)?)',  # Busca secuencias alfanuméricas largas
        r'(?:UNIDAD\s+)?([A-Z0-9]{4,}-[A-Z0-9]{4,})',    # Busca patrones con guión
        r'([A-Z]{2,4}-[A-Z0-9]{5,})',                    # Busca patrones como SM-X406BZAAPEO
        r'([A-Z0-9]{6,}(?:-[A-Z0-9]{2,})?)'              # Busca patrones como 7GTRCVA068
    ]

    for patron in patrones:
        match = re.search(patron, descripcion)
        if match:
            part_num = match.group(1)
            if len(part_num) >= 4:
                part_num = re.sub(r'[^\w\-]', '', part_num)
                return part_num.strip()

    return "No especificado"

def extract_category(descripcion):
    """Extrae la categoría del producto basado en la descripción."""
    desc_lower = descripcion.lower() if descripcion else ""
    for codigo, categoria in CATEGORIAS_OFICIALES.items():
        if categoria.lower() in desc_lower:
            return categoria
    return "OTROS"

def extract_brand(descripcion, marcas_list):
    """Extrae la marca del producto basado en la descripción."""
    desc_upper = descripcion.upper() if descripcion else ""
    for marca in marcas_list:
        if marca.upper() in desc_upper:
            return marca
    return "OTROS"

def get_all_zip_files_from_github(repo_url):
    """
    Obtiene todos los archivos ZIP del repositorio de GitHub usando la API.
    """
    try:
        # Convertir URL de GitHub a API
        api_url = repo_url.replace('github.com', 'api.github.com/repos')
        if not api_url.endswith('/contents'):
            api_url = api_url.rstrip('/') + '/contents'
        
        st.write(f"🔍 Buscando archivos en: {api_url}")
        
        headers = {'Accept': 'application/vnd.github.v3+json'}
        response = requests.get(api_url, headers=headers)
        response.raise_for_status()
        
        contents = response.json()
        zip_files = []
        
        for item in contents:
            if item.get('name', '').endswith('.zip'):
                zip_files.append({
                    'name': item['name'],
                    'download_url': item['download_url'],
                    'size': item.get('size', 0)
                })
        
        return zip_files
    
    except Exception as e:
        st.error(f"Error al obtener archivos del repositorio: {str(e)}")
        return []

def process_zip_file(zip_url, zip_name, progress_bar, status_text):
    """
    Procesa un archivo ZIP y extrae todos los JSON que contiene.
    """
    productos = []
    
    try:
        # Descargar ZIP
        status_text.text(f"📥 Descargando: {zip_name}")
        response = requests.get(zip_url)
        response.raise_for_status()
        
        # Procesar ZIP
        with zipfile.ZipFile(BytesIO(response.content)) as z:
            json_files = [f for f in z.namelist() if f.endswith('.json')]
            
            if not json_files:
                status_text.text(f"⚠️ No se encontraron JSON en {zip_name}")
                return productos
            
            total_json = len(json_files)
            for idx, json_file in enumerate(json_files):
                status_text.text(f"📄 Procesando {json_file} ({idx+1}/{total_json}) en {zip_name}")
                
                try:
                    with z.open(json_file) as f:
                        data = json.load(f)
                    
                    # Procesar según el formato del JSON
                    items = []
                    if isinstance(data, dict):
                        for key, value in data.items():
                            if isinstance(value, list):
                                items.extend(value)
                            elif isinstance(value, dict):
                                items.append(value)
                    elif isinstance(data, list):
                        items = data
                    
                    # Procesar cada item
                    for item in items:
                        if not isinstance(item, dict):
                            continue
                        
                        # Filtro estricto
                        fecha_pub = item.get('fecha_publicacion', '')
                        estado_ficha = item.get('estado_ficha', '')
                        estado_oferta = item.get('estado_oferta', '')
                        
                        # Verificar fecha en Junio 2026
                        if '06/2026' not in fecha_pub and '2026-06' not in fecha_pub:
                            continue
                        
                        # Verificar estados
                        if estado_ficha != "OFERTADA" or estado_oferta != "VIGENTE":
                            continue
                        
                        # Extraer datos
                        descripcion = item.get('descripcion', '')
                        part_number = extract_part_number(descripcion)
                        categoria = extract_category(descripcion)
                        marca = extract_brand(descripcion, MARCAS_COMPLETAS)
                        
                        producto = {
                            'ID_ProductoOfertado': item.get('ID_ProductoOfertado', ''),
                            'Part_Number': part_number,
                            'Categoria': categoria,
                            'Marca': marca,
                            'Descripcion': descripcion,
                            'Moneda': item.get('moneda', ''),
                            'Precio': float(item.get('precio', 0)),
                            'Fecha_Registro': item.get('fecha_registro', ''),
                            'Estado_Ficha': estado_ficha,
                            'Estado_Oferta': estado_oferta,
                            'Fecha_Adjudicacion': item.get('fecha_adjudicacion', ''),
                            'Fecha_Publicacion': fecha_pub,
                            'Motivo': item.get('motivo', ''),
                            'Justificacion': item.get('justificacion', ''),
                            'Puntaje': float(item.get('puntaje', 0)),
                            'Archivo_Origen': zip_name,
                            'JSON_Origen': json_file
                        }
                        productos.append(producto)
                        
                except Exception as e:
                    st.warning(f"⚠️ Error en {json_file}: {str(e)}")
                    continue
                
                # Actualizar progreso
                progress = (idx + 1) / total_json
                progress_bar.progress(progress)
        
        status_text.text(f"✅ Procesado: {zip_name} - {len(productos)} productos encontrados")
        
    except Exception as e:
        st.error(f"❌ Error al procesar {zip_name}: {str(e)}")
    
    return productos

def load_all_data_from_repo(repo_url):
    """
    Carga y procesa todos los datos de todos los ZIP en el repositorio.
    """
    all_productos = []
    
    # Obtener lista de archivos ZIP
    zip_files = get_all_zip_files_from_github(repo_url)
    
    if not zip_files:
        st.error("No se encontraron archivos ZIP en el repositorio.")
        return None
    
    st.info(f"📦 Se encontraron {len(zip_files)} archivos ZIP para procesar")
    
    # Crear contenedores para progreso
    progress_container = st.container()
    
    with progress_container:
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total_zips = len(zip_files)
        for idx, zip_info in enumerate(zip_files):
            status_text.text(f"🔄 Procesando ZIP {idx+1}/{total_zips}: {zip_info['name']}")
            
            productos = process_zip_file(
                zip_info['download_url'],
                zip_info['name'],
                progress_bar,
                status_text
            )
            
            all_productos.extend(productos)
            
            # Actualizar progreso general
            overall_progress = (idx + 1) / total_zips
            progress_bar.progress(overall_progress)
    
    status_text.text("✅ ¡Todos los archivos procesados!")
    
    if all_productos:
        return pd.DataFrame(all_productos)
    else:
        return None

def create_excel_report(df, filters):
    """Crea un archivo Excel con el dashboard y formato condicional."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Hoja de datos filtrados
        df_filtered = df.copy()
        df_filtered.to_excel(writer, sheet_name='Datos_Filtrados', index=False)

        # Hoja de resumen
        resumen = pd.DataFrame({
            'Métrica': [
                'Total de Fichas', 
                'Categorías', 
                'Marcas', 
                'Precio Promedio', 
                'Puntaje Promedio',
                'Archivos ZIP Procesados',
                'JSON Procesados'
            ],
            'Valor': [
                len(df_filtered),
                df_filtered['Categoria'].nunique(),
                df_filtered['Marca'].nunique(),
                f"${df_filtered['Precio'].mean():,.2f}",
                f"{df_filtered['Puntaje'].mean():.2f}",
                df_filtered['Archivo_Origen'].nunique(),
                df_filtered['JSON_Origen'].nunique()
            ]
        })
        resumen.to_excel(writer, sheet_name='Resumen', index=False)

        # Hoja de distribución por categoría
        categoria_counts = df_filtered['Categoria'].value_counts().reset_index()
        categoria_counts.columns = ['Categoria', 'Cantidad']
        categoria_counts.to_excel(writer, sheet_name='Dist_Categoria', index=False)

        # Hoja de distribución por marca
        marca_counts = df_filtered['Marca'].value_counts().reset_index()
        marca_counts.columns = ['Marca', 'Cantidad']
        marca_counts.to_excel(writer, sheet_name='Dist_Marca', index=False)

        # Hoja de distribución por archivo origen
        origen_counts = df_filtered['Archivo_Origen'].value_counts().reset_index()
        origen_counts.columns = ['Archivo_ZIP', 'Cantidad']
        origen_counts.to_excel(writer, sheet_name='Dist_Origen', index=False)

        # Aplicar formato condicional
        workbook = writer.book

        # Formato para la hoja de datos
        ws = writer.sheets['Datos_Filtrados']
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.column == 6:  # Columna de Precio
                    try:
                        if float(cell.value) > df_filtered['Precio'].quantile(0.75):
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    except:
                        pass
                elif cell.column == 14:  # Columna de Puntaje
                    try:
                        if float(cell.value) >= 85:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif float(cell.value) >= 70:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    except:
                        pass

    output.seek(0)
    return output

# --- INTERFAZ DE USUARIO ---
def main():
    st.markdown('<h1 class="main-header">📊 Dashboard de Fichas Técnicas - Junio 2026</h1>', unsafe_allow_html=True)
    
    # URL del repositorio
    repo_url = "https://github.com/StalinHA/EXTRAER_PARTNUMBE"
    
    # Contenedor para estado
    status_container = st.container()
    
    with status_container:
        st.markdown('<div class="info-box">📦 <b>Procesando todos los archivos ZIP del repositorio...</b></div>', unsafe_allow_html=True)
    
    # Cargar datos de todos los ZIP
    with st.spinner('🔄 Descargando y procesando todos los archivos ZIP del repositorio...'):
        df = load_all_data_from_repo(repo_url)

    if df is None or len(df) == 0:
        st.warning("⚠️ No se encontraron fichas que cumplan con los criterios (Junio 2026, OFERTADA + VIGENTE).")
        st.info("💡 Verifica que el repositorio contenga datos con las condiciones especificadas.")
        return

    # --- BARRA LATERAL DE FILTROS ---
    st.sidebar.header("🔍 Filtros")

    # Filtros
    estados = sorted(df['Estado_Ficha'].unique())
    estados_filter = st.sidebar.multiselect(
        "Estado de Ficha",
        options=estados,
        default=estados
    )

    categorias = sorted(df['Categoria'].unique())
    categorias_filter = st.sidebar.multiselect(
        "Categoría",
        options=categorias,
        default=categorias
    )

    marcas = sorted(df['Marca'].unique())
    marcas_filter = st.sidebar.multiselect(
        "Marca",
        options=marcas,
        default=marcas
    )

    # Rango de precios
    min_price = float(df['Precio'].min())
    max_price = float(df['Precio'].max())
    price_range = st.sidebar.slider(
        "Rango de Precio (USD)",
        min_value=min_price,
        max_value=max_price,
        value=(min_price, max_price),
        step=10.0,
        format="%.2f"
    )

    # Aplicar filtros
    df_filtered = df[
        (df['Estado_Ficha'].isin(estados_filter)) &
        (df['Categoria'].isin(categorias_filter)) &
        (df['Marca'].isin(marcas_filter)) &
        (df['Precio'] >= price_range[0]) &
        (df['Precio'] <= price_range[1])
    ]

    # --- MÉTRICAS PRINCIPALES ---
    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{len(df_filtered):,}</div>
            <div class="metric-label">Total de Fichas</div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{df_filtered['Categoria'].nunique()}</div>
            <div class="metric-label">Categorías</div>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{df_filtered['Marca'].nunique()}</div>
            <div class="metric-label">Marcas</div>
        </div>
        """, unsafe_allow_html=True)

    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">${df_filtered['Precio'].mean():,.2f}</div>
            <div class="metric-label">Precio Promedio</div>
        </div>
        """, unsafe_allow_html=True)

    with col5:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{df_filtered['Archivo_Origen'].nunique()}</div>
            <div class="metric-label">ZIP Procesados</div>
        </div>
        """, unsafe_allow_html=True)

    st.divider()

    # --- GRÁFICOS ---
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("📊 Distribución por Categoría")
        if not df_filtered.empty:
            fig = px.pie(
                df_filtered,
                names='Categoria',
                title='Fichas por Categoría',
                hole=0.3,
                color_discrete_sequence=px.colors.qualitative.Set3
            )
            fig.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay datos para mostrar")

    with col2:
        st.subheader("📈 Top 10 Marcas")
        if not df_filtered.empty:
            top_brands = df_filtered['Marca'].value_counts().head(10).reset_index()
            top_brands.columns = ['Marca', 'Cantidad']

            fig = px.bar(
                top_brands,
                x='Marca',
                y='Cantidad',
                title='Top 10 Marcas con más Fichas',
                color='Cantidad',
                color_continuous_scale='Blues',
                text='Cantidad'
            )
            fig.update_traces(textposition='outside')
            fig.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay datos para mostrar")

    # --- SEGUNDA FILA DE GRÁFICOS ---
    col3, col4 = st.columns(2)

    with col3:
        st.subheader("💰 Precio por Categoría")
        if not df_filtered.empty:
            price_by_cat = df_filtered.groupby('Categoria')['Precio'].mean().sort_values(ascending=False).reset_index()

            fig = px.bar(
                price_by_cat,
                x='Categoria',
                y='Precio',
                title='Precio Promedio por Categoría',
                color='Precio',
                color_continuous_scale='Reds',
                text='Precio'
            )
            fig.update_traces(texttemplate='$%{text:.2f}', textposition='outside')
            fig.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay datos para mostrar")

    with col4:
        st.subheader("📊 Resumen por Categoría")
        if not df_filtered.empty:
            summary = df_filtered.groupby('Categoria').agg(
                Cantidad=('Categoria', 'count'),
                Precio_Promedio=('Precio', 'mean'),
                Puntaje_Promedio=('Puntaje', 'mean')
            ).reset_index().sort_values('Cantidad', ascending=False)

            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=summary['Categoria'],
                y=summary['Cantidad'],
                name='Cantidad',
                marker_color='#3B82F6',
                yaxis='y'
            ))
            fig.add_trace(go.Scatter(
                x=summary['Categoria'],
                y=summary['Precio_Promedio'],
                name='Precio Promedio (USD)',
                marker_color='#EF4444',
                yaxis='y2',
                mode='lines+markers'
            ))

            fig.update_layout(
                title='Cantidad y Precio Promedio por Categoría',
                xaxis_title='Categoría',
                yaxis=dict(title='Cantidad', side='left'),
                yaxis2=dict(title='Precio Promedio (USD)', overlaying='y', side='right'),
                xaxis_tickangle=-45,
                legend=dict(x=0.01, y=0.99)
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay datos para mostrar")

    # --- TERCERA FILA: DISTRIBUCIÓN POR ARCHIVO ORIGEN ---
    st.subheader("📁 Distribución por Archivo ZIP")
    col5, col6 = st.columns(2)

    with col5:
        if not df_filtered.empty:
            origen_counts = df_filtered['Archivo_Origen'].value_counts().head(10).reset_index()
            origen_counts.columns = ['Archivo_ZIP', 'Cantidad']

            fig = px.bar(
                origen_counts,
                x='Archivo_ZIP',
                y='Cantidad',
                title='Top 10 Archivos ZIP con más Fichas',
                color='Cantidad',
                color_continuous_scale='Greens',
                text='Cantidad'
            )
            fig.update_traces(textposition='outside')
            fig.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay datos para mostrar")

    with col6:
        if not df_filtered.empty:
            # Gráfico de torta de origen
            fig = px.pie(
                df_filtered,
                names='Archivo_Origen',
                title='Distribución por Archivo ZIP',
                hole=0.3
            )
            fig.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay datos para mostrar")

    st.divider()

    # --- TABLA DE DATOS ---
    st.subheader("📋 Datos Detallados")

    # Mostrar número de registros
    st.caption(f"Mostrando {len(df_filtered):,} de {len(df):,} fichas que cumplen los filtros")

    # Columnas a mostrar
    display_cols = ['ID_ProductoOfertado', 'Part_Number', 'Categoria', 'Marca',
                    'Descripcion', 'Precio', 'Puntaje', 'Estado_Ficha', 'Estado_Oferta',
                    'Fecha_Publicacion', 'Archivo_Origen']

    # Crear tabla con estilo
    st.dataframe(
        df_filtered[display_cols],
        column_config={
            "ID_ProductoOfertado": st.column_config.TextColumn("ID Producto"),
            "Part_Number": st.column_config.TextColumn("Part Number"),
            "Categoria": st.column_config.TextColumn("Categoría"),
            "Marca": st.column_config.TextColumn("Marca"),
            "Descripcion": st.column_config.TextColumn("Descripción", width='large'),
            "Precio": st.column_config.NumberColumn("Precio (USD)", format="$%.2f"),
            "Puntaje": st.column_config.NumberColumn("Puntaje", format="%.2f"),
            "Estado_Ficha": st.column_config.TextColumn("Estado Ficha"),
            "Estado_Oferta": st.column_config.TextColumn("Estado Oferta"),
            "Fecha_Publicacion": st.column_config.TextColumn("Fecha Publicación"),
            "Archivo_Origen": st.column_config.TextColumn("Archivo ZIP"),
        },
        hide_index=True,
        use_container_width=True
    )

    # --- EXPORTAR A EXCEL ---
    st.divider()
    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        if st.button("📥 Exportar Dashboard a Excel", use_container_width=True, type="primary"):
            with st.spinner("Generando archivo Excel con dashboard..."):
                excel_data = create_excel_report(df_filtered, {})
                st.download_button(
                    label="📥 Descargar Excel",
                    data=excel_data,
                    file_name=f"dashboard_fichas_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success("✅ Excel generado correctamente!")

    # --- PIE DE PÁGINA ---
    st.divider()
    st.caption(f"📊 Dashboard generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Datos procesados desde GitHub")

if __name__ == "__main__":
    main()
